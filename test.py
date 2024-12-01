import pandas as pd
import sys
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
'''
前提
１）inputファイルはエクセルファイルとする
２）以下をインストールしておく（exeファイルを使用する場合は不要）
    import pandas as pd
    import sys
    import openpyxl
３）inputファイルの先頭行（１行目）はタイトル行とする
４）このソースコードに直接を書き込む必要があるので注意しておく
５）出力ファイルが生成される
６）直接書き込む（ハードコーディング）が必要なものについては要検討中
'''

'''
関数名：save_output
概要：Excelファイルを読み込む ハイパーリンクも保存してoutputファイルにinputファイルをコピーする
'''
def save_output(input_file,output_file):
    wb = openpyxl.load_workbook(input_file)
    wb.save(output_file) # エクセルのセーブ 一度セーブするとそのwbを更新することはできない
    print(f'>>>{input_file} から {output_file} へコピー完了しました')

'''
関数名：get_link_title
概要：linkと文字列を取得する
返り値：タイトルとリンクのリスト=list
'''
def get_link_title(ws):
    link = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink:
                hyperlink = cell.hyperlink.target
                text = cell.value
                temp_text_hype = []
                temp_text_hype.append(text)
                temp_text_hype.append(hyperlink)
                link.append(temp_text_hype)
    print(">>>タイトルとリンクのリストを取得しました")
    return link

'''
関数名：search_keyword
概要：指定の列においてキーワードがどれぐらい含まれているかを調べる
返り値：キーワードの個数=int
'''
def search_keyword(keyword,data):
    temp = data.split(',')# 文字列をカンマ区切りに分割する
    count = 0
    for k in temp:
        if keyword in k:
            count += 1
    return count

'''
関数名：keyword_num_write
概要：キーワードが含まれる個数分だけ新規の列に書き込む
返り値：なし
'''
def keyword_num_write(ws):
    row_index = 1 # 読み込み行インデックス
    no_keyword = 1 # キーワードがない時は1を代入する
    # 特定の列のデータを一行ずつ処理(1行目は読み飛ばす)
    for row in ws.iter_rows(min_row=2, min_col=target_column, max_col=target_column):
        for cell in row:
            row_index += 1
            keyword_num = search_keyword(keyword,cell.value)
            if keyword_num: #キーワードが存在する場合
                ws.cell(row=row_index, column=keyword_num_column, value=keyword_num)
            else: #キーワードが存在しない場合
                ws.cell(row=row_index, column=keyword_num_column, value=no_keyword)
    print(">>>キーワードが含まれる個数を新規の列に書き込みました")

'''
関数名：first_save
概要：キーワードの個数だけ行を複製してをセーブする
返り値：なし
'''
def first_save(ws):
    # 新しいワークブックを作成し、シートを追加　
    new_wb = Workbook()
    new_ws = new_wb.active

    # 元のワークシートのデータを処理　
    new_ws.append(item_row) # input_fileの項目行(１行目)の追加
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        count = row[keyword_num_column -1]  # キーワードの個数を書く列の値を取得
        if isinstance(count, int):  # キーワードの個数が整数であることを確認
            for _ in range(count):  # キーワードの個数だけ行を複製
                new_ws.append(row)
    
    # エクセルのセーブ 一度セーブするとそのwbを更新することはできない
    new_wb.save(output_file)

'''
関数名：write_df_keyword
概要：キーワードを該当の列に記載してさらに分割してものも記載する
返り値：なし
'''
def write_df_keyword(df):
    # 特定の列で同じ値を持つ行を取得（例: 'title_column_name'列が'specific_value'の行を取得）
    for tt in range(len(title)):
        specific_value = title[tt]
        # データフレームの列を object 型に変換することで、異なるデータ型を許容できるようにする。
        df[item_row[keyword_splitwrite_column - 1]] = df[item_row[keyword_splitwrite_column - 1]].astype('object')
        df[item_row[keyword_splitcenterwrite_column - 1]] = df[item_row[keyword_splitcenterwrite_column - 1]].astype('object')
        df[item_row[keyword_splitlastwrite_column - 1]] = df[item_row[keyword_splitlastwrite_column - 1]].astype('object')

        filtered_df = df[df[title_column_name] == specific_value].copy()
        row_indexs = filtered_df.index #filtered_dfの行番号を取得

        no_key_count = 0
        no_key_write = 'NO'
        num_conut = 1
        for i in row_indexs:
            temp = filtered_df.at[i,item_row[target_column - 1]].split(',')# 文字列をカンマ区切りに分割する
            count = 0
            for k in temp:
                if keyword in k:
                    count += 1
                    if num_conut == count:
                        filtered_df.at[i,item_row[keyword_splitwrite_column - 1]] = k
                        split_text = k.split('_') #アンダースコアで文字列を分割する
                        filtered_df.at[i,item_row[keyword_splitcenterwrite_column - 1]] = split_text[1] # キーワードの1つを_(アンダースコア)で分割した２番目
                        split_text[2] = split_text[2].lstrip('0') #先頭の"0"を消去する
                        split_text[2] = int(split_text[2])
                        filtered_df.at[i,item_row[keyword_splitlastwrite_column - 1]] = split_text[2] # キーワードの1つを_(アンダースコア)で分割した３番目
                        num_conut += 1
                        break
                    else:
                        pass
                else:
                    no_key_count +=1
                    if len(temp) == no_key_count:
                        filtered_df.at[i,item_row[keyword_splitwrite_column - 1]] = no_key_write
                        filtered_df.at[i,item_row[keyword_splitcenterwrite_column - 1]] = no_key_write # キーワードの1つを_(アンダースコア)で分割した２番目
                        filtered_df.at[i,item_row[keyword_splitlastwrite_column - 1]] = no_key_write # キーワードの1つを_(アンダースコア)で分割した３番目
                    else:
                        pass
        
        df.update(filtered_df) # フィルター結果を元のデータフレームに反映

'''
関数名：write_hyperlink
概要：タイトルのハイパーリンクを追加する
返り値：なし
'''
def append_hyperlink(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=title_column, max_col=title_column):
        for cell in row:
            if cell.value in title:
                for sublist in hyperlink_title:
                    if cell.value == sublist[0]:
                        cell.hyperlink = sublist[1]
                        cell.font = Font(color="0000FF", underline="single")
                    else:
                        pass
            else:
                pass

'''
関数名：arrange_excel
概要：不要なデータを削除して列データを入れ替える
返り値：なし
'''
def arrange_excel(ws):
    ws.delete_cols(keyword_num_column) # キーワードの個数を書く列番号(Aであれば1、Cなら3)を削除
    ws.delete_cols(target_column) # キーワードの文字列が含まれる列番号(Aであれば1、Cなら3)を削除
    

    # 列のデータを一時的に保存
    data_col1 = [cell for cell in ws.iter_cols(min_col=col1, max_col=col1, values_only=True)]
    data_col2 = [cell for cell in ws.iter_cols(min_col=col2, max_col=col2, values_only=True)]

    # 列のデータを入れ替え
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=col1).value = data_col2[0][row-1]
        ws.cell(row=row, column=col2).value = data_col1[0][row-1]

if __name__ == '__main__':
    # データの読み込みとコピー　
    input_file = sys.argv[1]
    output_file = 'output_file.xlsx' # 出力エクセルファイル名　★直接書き込み記載あり
    redirect_file = 'output.txt' # ログテキストファイル名　★直接書き込み記載あり

    print("ツールの使用開始")
    # テキストファイルにリダイレクトするためのファイルオブジェクトを作成
    with open(redirect_file, 'w') as f:
        # sys.stdout をテキストファイルにリダイレクト
        sys.stdout = f

        item_row_num = 1 # input_fileの項目行(番号やタイトルが書かれた行）（先頭行であってほしい)　★直接書き込み記載あり
        save_output(input_file,output_file)

        # # Excelファイルを読み込む outputファイル
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # output_fileの項目行(１行目)の読み込み 
        item_row = next(ws.iter_rows(min_row=item_row_num, max_row=item_row_num, values_only=True))

        # タイトルとハイパーリンクの取得　
        hyperlink_title = get_link_title(ws) #タイトルとハイパーリンクのリスト
        title = [sublist[0] for sublist in hyperlink_title] #タイトルのみのリスト

        # キーワードの個数の取得と貼り付け　
        keyword = 'PUD' # キーワードの文字列　★直接書き込み記載あり
        target_column = 3 # キーワードの文字列が含まれる列番号(Aであれば1、Cなら3)を指定する　★直接書き込み記載あり⇒後にarrange_excel関数で削除される
        keyword_num_column = 4 # キーワードの個数を書く列番号(Aであれば1、Cなら3)を指定する　★直接書き込み記載あり⇒後にarrange_excel関数で削除される
        keyword_splitwrite_column = 5 # キーワードの1つ記載する列番号(Aであれば1、Cなら3)を指定する　★直接書き込み記載あり
        keyword_splitcenterwrite_column = 6 # キーワードの1つを_(アンダースコア)で分割した２番目を書く列番号(Aであれば1、Cなら3)を指定する　★直接書き込み記載あり
        keyword_splitlastwrite_column = 7 # キーワードの1つを_(アンダースコア)で分割した３番目を書く列番号(Aであれば1、Cなら3)を指定する　★直接書き込み記載あり
        title_column = 2 # タイトル列番号(Aであれば1、Cなら3)を指定する　★直接書き込み記載あり
        title_column_name = item_row[1]  # タイトル列名 >>>列インデックス(Aであれば0、Cなら2)を指定する(1 = 文献リスト) ★直接書き込み記載あり

        keyword_num_write(ws) #キーワードが含まれる個数分だけ新規の列に書き込む
        first_save(ws) # キーワードの個数だけ行を複製してをセーブする

        # outputをpandasで読み込む
        df = pd.read_excel(output_file)
        df = df.fillna(0) #空のデータを0で埋める

        write_df_keyword(df) #キーワードを該当の列に記載してさらに分割してものも記載する
        df.to_excel(output_file, index=False) #dfをoutputファイルに保存する

        # # Excelファイルを読み込む outputファイル
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        append_hyperlink(ws) #タイトルのハイパーリンクを追加する

        # 列Cと列Dを入れ替える（列番号は3と4）
        col1 = 3 #入れ替える列番号(Aであれば1、Cなら3)を指定する★直接書き込み記載あり
        col2 = 4 #入れ替える列番号(Aであれば1、Cなら3)を指定する★直接書き込み記載あり
        arrange_excel(ws) #不要なデータを削除して列データを入れ替える

        # エクセルのセーブ 一度セーブするとそのwbを更新することはできない
        wb.save(output_file)

        # リダイレクトを元に戻す
        sys.stdout = sys.__stdout__
    print(f"ツールの完了と出力ファイル作成({output_file})、ログファイル({redirect_file})が完了しました。")
